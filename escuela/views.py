from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
import unicodedata

import json

from .models import ClaseEscuela, Alumno, Asistencia, DomingoExcluido, GRADO_CHOICES
from miembros.models import Hermano
from datetime import date, timedelta


def _quitar_acentos(texto):
    """Quita acentos de un texto para busqueda flexible"""
    if not texto:
        return ''
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    ).lower()


def _parsear_fecha_selects(post_data):
    """Parsea fecha de nacimiento desde 3 selects (dia, mes, anio)"""
    dia = post_data.get('nacimiento_dia', '').strip()
    mes = post_data.get('nacimiento_mes', '').strip()
    anio = post_data.get('nacimiento_anio', '').strip()

    if dia and mes and anio:
        try:
            from datetime import date
            return date(int(anio), int(mes), int(dia))
        except (ValueError, TypeError):
            return None
    return None


def acceso(request):
    """
    Pantalla de acceso: el maestro ingresa el codigo de su clase
    """
    # Si ya tiene acceso, ir a lista
    if request.session.get('escuela_clase_id'):
        return redirect('escuela:lista_alumnos')

    error = None
    clases = ClaseEscuela.objects.filter(activo=True)

    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip().upper()

        if not codigo:
            error = 'Ingresa el código de acceso.'
        else:
            # Buscar clase por código
            try:
                clase = ClaseEscuela.objects.get(codigo_acceso__iexact=codigo, activo=True)
                request.session['escuela_clase_id'] = clase.id
                return redirect('escuela:lista_alumnos')
            except ClaseEscuela.DoesNotExist:
                error = 'Código de acceso incorrecto.'

    return render(request, 'escuela/acceso.html', {
        'error': error,
        'clases': clases,
    })


def lista_alumnos(request):
    """
    Lista de alumnos de la clase del maestro
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)
    alumnos = list(Alumno.objects.filter(clase=clase, activo=True).order_by('apellidos', 'nombres'))

    # Detectar cumpleañeros de los ultimos 7 dias
    hoy = date.today()
    cumple_ids = set()
    for alumno in alumnos:
        if alumno.fecha_nacimiento:
            for i in range(7):
                dia = hoy - timedelta(days=i)
                if dia.month == alumno.fecha_nacimiento.month and dia.day == alumno.fecha_nacimiento.day:
                    cumple_ids.add(alumno.id)
                    break

    # Ordenar: cumpleañeros primero, luego alfabetico
    alumnos.sort(key=lambda a: (a.id not in cumple_ids, a.apellidos.lower(), a.nombres.lower()))

    return render(request, 'escuela/lista_alumnos.html', {
        'clase': clase,
        'alumnos': alumnos,
        'cumple_ids': cumple_ids,
    })


def alumno_crear(request):
    """
    Crear nuevo alumno - optimizado para carga rapida desde celular
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)

    if request.method == 'POST':
        apellidos = request.POST.get('apellidos', '').strip()
        nombres = request.POST.get('nombres', '').strip()
        grado_nivel = request.POST.get('grado_nivel', '').strip()
        observaciones = request.POST.get('observaciones', '').strip()

        fecha_nacimiento = _parsear_fecha_selects(request.POST)

        if not apellidos or not nombres:
            messages.error(request, 'Apellidos y nombres son obligatorios.')
        else:
            # Verificar duplicado en cualquier clase
            duplicado = Alumno.objects.filter(
                apellidos__iexact=apellidos,
                nombres__iexact=nombres,
                activo=True,
            ).first()

            if duplicado and 'confirmar_duplicado' not in request.POST:
                # Hay duplicado y no confirmo: mostrar advertencia
                return render(request, 'escuela/alumno_form.html', {
                    'clase': clase,
                    'grado_choices': GRADO_CHOICES,
                    'dias': range(1, 32),
                    'es_nuevo': True,
                    'duplicado': duplicado,
                    'form_data': request.POST,
                })

            Alumno.objects.create(
                apellidos=apellidos,
                nombres=nombres,
                fecha_nacimiento=fecha_nacimiento,
                grado_nivel=grado_nivel,
                clase=clase,
                observaciones=observaciones,
            )

            # Si pidio "guardar y cargar otro"
            if 'guardar_otro' in request.POST:
                messages.success(request, f'{nombres} {apellidos} registrado correctamente.')
                return redirect('escuela:alumno_crear')

            messages.success(request, f'{nombres} {apellidos} registrado correctamente.')
            return redirect('escuela:lista_alumnos')

    return render(request, 'escuela/alumno_form.html', {
        'clase': clase,
        'grado_choices': GRADO_CHOICES,
        'dias': range(1, 32),
        'es_nuevo': True,
    })


def alumno_editar(request, alumno_id):
    """
    Editar alumno existente
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)
    alumno = get_object_or_404(Alumno, id=alumno_id, clase=clase)

    if request.method == 'POST':
        apellidos = request.POST.get('apellidos', '').strip()
        nombres = request.POST.get('nombres', '').strip()
        grado_nivel = request.POST.get('grado_nivel', '').strip()
        observaciones = request.POST.get('observaciones', '').strip()

        fecha_nacimiento = _parsear_fecha_selects(request.POST)

        if not apellidos or not nombres:
            messages.error(request, 'Apellidos y nombres son obligatorios.')
        else:
            alumno.apellidos = apellidos
            alumno.nombres = nombres
            alumno.fecha_nacimiento = fecha_nacimiento
            alumno.grado_nivel = grado_nivel
            alumno.observaciones = observaciones
            alumno.save()

            messages.success(request, f'Datos de {nombres} {apellidos} actualizados.')
            return redirect('escuela:lista_alumnos')

    return render(request, 'escuela/alumno_form.html', {
        'clase': clase,
        'alumno': alumno,
        'grado_choices': GRADO_CHOICES,
        'dias': range(1, 32),
        'es_nuevo': False,
    })


def alumno_eliminar(request, alumno_id):
    """
    Eliminar (desactivar) alumno
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)
    alumno = get_object_or_404(Alumno, id=alumno_id, clase=clase)

    if request.method == 'POST':
        alumno.activo = False
        alumno.save()
        messages.success(request, f'{alumno.nombre_completo()} fue eliminado de la lista.')
        return redirect('escuela:lista_alumnos')

    return render(request, 'escuela/alumno_eliminar.html', {
        'clase': clase,
        'alumno': alumno,
    })


def buscar_hermanos(request):
    """
    API para autocompletado: busca hermanos por apellido/nombre
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return JsonResponse([], safe=False)

    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse([], safe=False)

    q_norm = _quitar_acentos(q)

    # Buscar en memoria con comparacion sin acentos
    todos = Hermano.objects.filter(activo=True).order_by('nombres', 'apellidos')
    resultados = []
    for h in todos:
        if (q_norm in _quitar_acentos(h.apellidos)
                or q_norm in _quitar_acentos(h.nombres)):
            resultados.append({
                'apellidos': h.apellidos,
                'nombres': h.nombres,
                'fecha_nacimiento': h.fecha_nacimiento.strftime('%Y-%m-%d') if h.fecha_nacimiento else '',
                'edad': h.edad(),
            })
            if len(resultados) >= 10:
                break

    return JsonResponse(resultados, safe=False)


def _domingo_mas_reciente():
    """Devuelve el domingo mas reciente (hoy si es domingo)"""
    hoy = date.today()
    dias_desde_domingo = hoy.weekday() + 1  # lunes=0 -> +1=1, domingo=6 -> +1=7
    if dias_desde_domingo == 7:
        return hoy
    return hoy - timedelta(days=dias_desde_domingo)


def _es_cumple_semana(fecha_nacimiento, fecha_domingo):
    """Verifica si el cumpleaños cae en la semana del domingo (lunes a domingo)"""
    if not fecha_nacimiento:
        return False
    lunes = fecha_domingo - timedelta(days=6)
    for i in range(7):
        dia = lunes + timedelta(days=i)
        if dia.month == fecha_nacimiento.month and dia.day == fecha_nacimiento.day:
            return True
    return False


def _domingos_disponibles(n=8):
    """Devuelve el proximo domingo + los ultimos N domingos, excluyendo domingos sin clases"""
    domingo = _domingo_mas_reciente()
    proximo = domingo + timedelta(weeks=1)
    candidatos = [proximo]
    for i in range(n):
        candidatos.append(domingo - timedelta(weeks=i))

    excluidos = set(DomingoExcluido.objects.values_list('fecha', flat=True))
    return [d for d in candidatos if d not in excluidos]


def asistencia_inicio(request):
    """
    Seleccionar domingo para pasar asistencia
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)
    total_alumnos = Alumno.objects.filter(clase=clase, activo=True).count()

    if total_alumnos < 5:
        messages.warning(request, 'Se necesitan al menos 5 alumnos para habilitar la asistencia.')
        return redirect('escuela:lista_alumnos')

    domingos = _domingos_disponibles()

    # Marcar domingos que ya tienen asistencia registrada
    domingos_info = []
    for domingo in domingos:
        registros = Asistencia.objects.filter(clase=clase, fecha=domingo).count()
        domingos_info.append({
            'fecha': domingo,
            'registros': registros,
            'completa': registros >= total_alumnos,
        })

    return render(request, 'escuela/asistencia_inicio.html', {
        'clase': clase,
        'domingos': domingos_info,
        'total_alumnos': total_alumnos,
    })


def asistencia_pasar(request, fecha_str):
    """
    Pasar lista: muestra alumno por alumno
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)

    try:
        from datetime import datetime
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return redirect('escuela:asistencia_inicio')

    # Bloquear domingos excluidos
    if DomingoExcluido.objects.filter(fecha=fecha).exists():
        messages.warning(request, 'Este domingo no hay clases de escuela bíblica.')
        return redirect('escuela:asistencia_inicio')

    alumnos = list(Alumno.objects.filter(clase=clase, activo=True).order_by('apellidos', 'nombres'))

    if len(alumnos) < 5:
        return redirect('escuela:lista_alumnos')

    # Obtener asistencias ya registradas para esta fecha
    asistencias_existentes = {
        a.alumno_id: a.presente
        for a in Asistencia.objects.filter(clase=clase, fecha=fecha)
    }

    # Armar lista con estado
    alumnos_data = []
    for alumno in alumnos:
        alumnos_data.append({
            'id': alumno.id,
            'nombres': alumno.nombres,
            'apellidos': alumno.apellidos,
            'edad': alumno.edad(),
            'estado': asistencias_existentes.get(alumno.id),  # None, True o False
            'cumple': _es_cumple_semana(alumno.fecha_nacimiento, fecha),
        })

    return render(request, 'escuela/asistencia_pasar.html', {
        'clase': clase,
        'fecha': fecha,
        'alumnos_json': json.dumps(alumnos_data),
        'total': len(alumnos_data),
        'ya_registrados': len(asistencias_existentes),
    })


def asistencia_registrar(request):
    """
    API: registrar asistencia de un alumno (llamado via fetch)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)

    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return JsonResponse({'error': 'Sin acceso'}, status=403)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON invalido'}, status=400)

    alumno_id = data.get('alumno_id')
    fecha_str = data.get('fecha')
    presente = data.get('presente')

    if alumno_id is None or fecha_str is None or presente is None:
        return JsonResponse({'error': 'Datos incompletos'}, status=400)

    try:
        from datetime import datetime
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Fecha invalida'}, status=400)

    # Bloquear domingos excluidos
    if DomingoExcluido.objects.filter(fecha=fecha).exists():
        return JsonResponse({'error': 'Domingo sin clases'}, status=400)

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)
    alumno = get_object_or_404(Alumno, id=alumno_id, clase=clase)

    # Crear o actualizar
    asistencia, created = Asistencia.objects.update_or_create(
        alumno=alumno,
        fecha=fecha,
        defaults={
            'clase': clase,
            'presente': presente,
        }
    )

    return JsonResponse({'ok': True, 'created': created})


def asistencia_historial(request):
    """
    Historial de asistencias de la clase
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)

    # Ultimos 8 domingos con asistencia
    fechas = (Asistencia.objects.filter(clase=clase)
              .values_list('fecha', flat=True)
              .distinct()
              .order_by('-fecha')[:8])

    historial = []
    for fecha in fechas:
        registros = Asistencia.objects.filter(clase=clase, fecha=fecha)
        presentes = registros.filter(presente=True).count()
        ausentes = registros.filter(presente=False).count()
        total = presentes + ausentes
        historial.append({
            'fecha': fecha,
            'presentes': presentes,
            'ausentes': ausentes,
            'total': total,
            'porcentaje': round(presentes * 100 / total) if total > 0 else 0,
        })

    return render(request, 'escuela/asistencia_historial.html', {
        'clase': clase,
        'historial': historial,
    })


def asistencia_detalle(request, fecha_str):
    """
    Detalle de asistencia de un domingo: lista de presentes y ausentes
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)

    try:
        from datetime import datetime
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return redirect('escuela:asistencia_historial')

    registros = list(Asistencia.objects.filter(clase=clase, fecha=fecha)
                     .select_related('alumno')
                     .order_by('alumno__apellidos', 'alumno__nombres'))

    # Marcar cumpleañeros
    cumple_ids = set()
    for r in registros:
        if _es_cumple_semana(r.alumno.fecha_nacimiento, fecha):
            cumple_ids.add(r.alumno.id)

    # Ordenar: cumpleañeros primero, luego alfabetico
    registros.sort(key=lambda r: (r.alumno.id not in cumple_ids, r.alumno.apellidos.lower(), r.alumno.nombres.lower()))

    presentes = [r for r in registros if r.presente]
    ausentes = [r for r in registros if not r.presente]
    total = len(registros)
    porcentaje = round(len(presentes) * 100 / total) if total > 0 else 0

    return render(request, 'escuela/asistencia_detalle.html', {
        'clase': clase,
        'fecha': fecha,
        'presentes': presentes,
        'ausentes': ausentes,
        'total': total,
        'porcentaje': porcentaje,
        'cumple_ids': cumple_ids,
    })


def exportar_excel(request):
    """
    Exportar planilla Excel con alumnos y asistencia por domingo
    """
    clase_id = request.session.get('escuela_clase_id')
    if not clase_id:
        return redirect('escuela:acceso')

    clase = get_object_or_404(ClaseEscuela, id=clase_id, activo=True)
    alumnos = Alumno.objects.filter(clase=clase, activo=True).order_by('apellidos', 'nombres')

    # Obtener todos los domingos con asistencia, ordenados de mas viejo a mas nuevo
    fechas = list(
        Asistencia.objects.filter(clase=clase)
        .values_list('fecha', flat=True)
        .distinct()
        .order_by('fecha')
    )

    # Armar diccionario de asistencias: {(alumno_id, fecha): presente}
    asistencias = {}
    for a in Asistencia.objects.filter(clase=clase):
        asistencias[(a.alumno_id, a.fecha)] = a.presente

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = clase.nombre

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    fecha_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    presente_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    ausente_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(fechas))
    title_cell = ws.cell(row=1, column=1,
                         value=f'{clase.emoji} {clase.nombre} - Escuela Bíblica 2025')
    title_cell.font = Font(bold=True, size=14, color='2E7D32')
    title_cell.alignment = Alignment(horizontal='center')

    # Subtitulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3 + len(fechas))
    sub_cell = ws.cell(row=2, column=1,
                       value=f'Maestros: {clase.maestros} | Rango: {clase.rango_edad()}')
    sub_cell.font = Font(size=10, color='666666')
    sub_cell.alignment = Alignment(horizontal='center')

    # Encabezados
    row = 4
    headers = ['Apellidos', 'Nombres', 'Edad']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Encabezados de fechas
    for i, fecha in enumerate(fechas):
        cell = ws.cell(row=row, column=4 + i, value=fecha.strftime('%d/%m'))
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = fecha_fill
        cell.alignment = center
        cell.border = thin_border

    # Datos
    for idx, alumno in enumerate(alumnos):
        r = row + 1 + idx
        ws.cell(row=r, column=1, value=alumno.apellidos).border = thin_border
        ws.cell(row=r, column=2, value=alumno.nombres).border = thin_border

        edad_cell = ws.cell(row=r, column=3, value=alumno.edad())
        edad_cell.alignment = center
        edad_cell.border = thin_border

        for i, fecha in enumerate(fechas):
            key = (alumno.id, fecha)
            cell = ws.cell(row=r, column=4 + i)
            cell.alignment = center
            cell.border = thin_border

            if key in asistencias:
                if asistencias[key]:
                    cell.value = 'P'
                    cell.fill = presente_fill
                    cell.font = Font(bold=True, color='2E7D32')
                else:
                    cell.value = 'A'
                    cell.fill = ausente_fill
                    cell.font = Font(bold=True, color='C62828')
            else:
                cell.value = '-'
                cell.font = Font(color='AAAAAA')

    # Fila de totales
    if fechas and alumnos:
        r_total = row + 1 + len(alumnos)
        ws.cell(row=r_total, column=1, value='TOTAL PRESENTES').font = Font(bold=True)
        ws.cell(row=r_total, column=1).border = thin_border
        ws.cell(row=r_total, column=2).border = thin_border
        ws.cell(row=r_total, column=3).border = thin_border

        for i, fecha in enumerate(fechas):
            presentes = sum(1 for a in alumnos if asistencias.get((a.id, fecha)) is True)
            total_reg = sum(1 for a in alumnos if (a.id, fecha) in asistencias)
            cell = ws.cell(row=r_total, column=4 + i)
            cell.value = f'{presentes}/{total_reg}' if total_reg > 0 else '-'
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = thin_border

    # Anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    for i in range(len(fechas)):
        col_letter = openpyxl.utils.get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 8

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f'escuela_{clase.nombre.lower().replace(" ", "_")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


def cerrar_sesion(request):
    """Cerrar sesion de escuela"""
    request.session.pop('escuela_clase_id', None)
    return redirect('escuela:acceso')
